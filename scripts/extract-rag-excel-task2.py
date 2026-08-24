from __future__ import annotations

import argparse
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PARSER_VERSION = "rag-excel-task2/2.0.0"
DIMENSIONS = ("TR", "CC", "LR", "GRA")
SHEET_ALIASES = {
    "global": "整篇四维",
    "paragraph": "段落四维评价",
    "sentence": "句子",
    "opinions": "段落评价_看法",
    "teaching": "Teaching_note",
    "essay_paragraphs": "学生原文段_mark",
    "expansions": "段级合并_Expansion",
    "model": "MODEL范文",
    "model_analysis": "MODEL为什么好",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Normalize IELTS Task 2 RAG Excel workbooks into a lossless manifest.")
    parser.add_argument("--input-dir", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--question-overrides", type=Path)
    parser.add_argument("--holdout-rate", type=float, default=0.20)
    parser.add_argument("--summary-only", action="store_true")
    return parser.parse_args()


def clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").replace("\xa0", " ")
    text = unicodedata.normalize("NFKC", text)
    text = "\n".join(re.sub(r"[ \t]+", " ", line).strip() for line in text.split("\n"))
    text = re.sub(r"\n{3,}", "\n\n", text).strip()
    return text or None


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sha256_text(text: str) -> str:
    return hashlib.sha256(re.sub(r"\s+", " ", text).strip().lower().encode("utf-8")).hexdigest()


def english_word_count(text: str) -> int:
    return len(re.findall(r"[A-Za-z]+(?:['’-][A-Za-z]+)?", text))


def unique(items: list[str]) -> list[str]:
    return list(dict.fromkeys(item for item in items if item))


def rows_for_sheet(workbook, sheet_name: str) -> list[list[str | None]]:
    if sheet_name not in workbook.sheetnames:
        return []
    result: list[list[str | None]] = []
    for row in workbook[sheet_name].iter_rows(values_only=True):
        values = [clean(value) for value in row]
        while values and values[-1] is None:
            values.pop()
        if any(value is not None for value in values):
            result.append(values)
    return result


def row_objects(rows: list[list[str | None]], sheet_name: str) -> list[dict[str, Any]]:
    if not rows:
        return []
    headers = [value or f"column_{idx + 1}" for idx, value in enumerate(rows[0])]
    result = []
    for excel_row, values in enumerate(rows[1:], start=2):
        record: dict[str, Any] = {
            headers[idx]: values[idx] if idx < len(values) else None
            for idx in range(len(headers))
        }
        record["__sourceSheet"] = sheet_name
        record["__sourceRow"] = excel_row
        result.append(record)
    return result


def source_cell(sheet: str, row: int, column: str) -> dict[str, Any]:
    return {"sheet": sheet, "row": row, "column": column}


def extract_sid_refs(text: str | None) -> list[str]:
    if not text:
        return []
    # Preserve optional workbook/version prefixes. They are essential in files
    # such as xc18, where a-s01 and b-s01 are different student drafts.
    refs = re.findall(
        # Prefix segments may be workbook/version identifiers (xc18-a-s01),
        # but a preceding SID in a range (s01-s02) must not be swallowed as
        # the prefix of the second SID.
        r"(?<![A-Za-z0-9])(?:(?!s\d{1,3}-)[A-Za-z_][A-Za-z0-9_]*-){0,2}s\d{1,3}(?![A-Za-z0-9])",
        text,
        flags=re.I,
    )
    return unique([ref.lower() for ref in refs])


def sid_number(sid: str) -> int | None:
    match = re.search(r"s(\d{1,3})$", sid, flags=re.I)
    return int(match.group(1)) if match else None


def sid_family(sid: str) -> str:
    return re.sub(r"s\d{1,3}$", "", sid.lower())


def sid_version(sid: str) -> str:
    """Return a local draft family such as ``a`` or ``b`` when present."""
    family = sid_family(sid).rstrip("-")
    if not family:
        return ""
    tail = family.rsplit("-", 1)[-1]
    return tail if re.fullmatch(r"[a-z]", tail, flags=re.I) else ""


def parse_band(text: str | None) -> float | None:
    if not text:
        return None
    match = re.search(r"(?:预估分数|综合预估|overall(?:\s+band)?)[^0-9]{0,12}([0-9](?:\.5)?)", text, flags=re.I)
    if not match:
        return None
    value = float(match.group(1))
    return value if 0 <= value <= 9 else None


def infer_finding_kind(content: str) -> str:
    if re.search(r"矛盾|冲突|不一致|无法协调|相反", content):
        return "CONTRADICTION"
    if re.search(r"建议|应当|应该|可改|可以改|更适合|需要改|宜写|宜改", content):
        return "RECOMMENDATION"
    if re.search(r"展开|论证|解释|证明|因果|支撑", content):
        return "DEVELOPMENT"
    if re.search(r"问题|错误|不足|没有|未能|无法|不自然|不准确|过强|偏题", content):
        return "PROBLEM"
    return "JUDGEMENT"


def split_feedback_findings(text: str | None, default_scope: str) -> list[dict[str, Any]]:
    if not text:
        return []
    # Keep offsets in the complete dimension feedback. Semicolons are useful
    # boundaries in these workbooks because teachers often place one diagnosis
    # or recommendation in each clause.
    findings = []
    evidence_appendix = False
    pattern = re.compile(r"[^。！？；;\n]+(?:[。！？；;]|\n+|$)", flags=re.S)
    for raw_match in pattern.finditer(text):
        raw = raw_match.group(0)
        content = raw.strip()
        if not content:
            continue
        leading = len(raw) - len(raw.lstrip())
        trailing = len(raw.rstrip())
        refs = extract_sid_refs(content)
        if "完整原句定位" in content:
            evidence_appendix = True
        finding_scope = default_scope
        if len(refs) == 1:
            finding_scope = "SENTENCE"
        elif len(refs) > 1:
            finding_scope = "CROSS_SCOPE"
        findings.append({
            "ordinal": len(findings),
            "scope": finding_scope,
            "kind": "EVIDENCE_QUOTE" if evidence_appendix else infer_finding_kind(content),
            "content": content,
            "feedbackStartOffset": raw_match.start() + leading,
            "feedbackEndOffset": raw_match.start() + trailing,
            "evidenceSids": refs,
            "confidence": 1.0 if refs else 0.85,
            "semanticStatus": "DETERMINISTIC",
        })
    if not findings:
        findings.append({
            "ordinal": 0,
            "scope": default_scope,
            "kind": infer_finding_kind(text),
            "content": text,
            "feedbackStartOffset": 0,
            "feedbackEndOffset": len(text),
            "evidenceSids": extract_sid_refs(text),
            "confidence": 0.8,
            "semanticStatus": "DETERMINISTIC",
        })
    return findings


def parse_dimension_feedback(
    text: str | None,
    source: dict[str, Any] | None = None,
    default_scope: str = "SENTENCE",
) -> list[dict[str, Any]]:
    if not text:
        return []
    header_pattern = re.compile(
        r"(?:^|\n)\s*(?:\d+[.、]\s*)?\[(TR|CC|LR|GRA)\](?:\[([^\]]+)\])?\s*"
        r"(?:\[([^\]]+)\]|[（(]([^）)]+)[）)])?\s*",
        flags=re.I,
    )
    matches = list(header_pattern.finditer(text))
    results = []
    for index, match in enumerate(matches):
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        feedback = clean(text[match.end():end]) or ""
        tags_raw = match.group(3) or match.group(4) or ""
        results.append({
            "dimension": match.group(1).upper(),
            "status": match.group(2).upper() if match.group(2) else None,
            "tags": unique([part.strip().upper() for part in re.split(r"[,，、]", tags_raw)]),
            "feedback": feedback,
            "evidenceSids": extract_sid_refs(feedback),
            "findings": split_feedback_findings(feedback, default_scope),
            "source": source,
        })
    return results


def parse_word_annotations(text: str | None, source: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    if not text:
        return []
    normalized = text.replace("（", "(").replace("）", ")")
    pattern = re.compile(
        r"\[([^\]]+)\]\s*\(所在句\s*:\s*(.*?)\)\s*\n(.*?)(?=\n\n\[[^\]]+\]\s*\(所在句\s*:|\Z)",
        flags=re.S,
    )
    results = []
    for index, match in enumerate(pattern.finditer(normalized), start=1):
        anchor = clean(match.group(1))
        sentence = clean(match.group(2))
        feedback = clean(match.group(3))
        if feedback:
            results.append({
                "index": index,
                "anchorText": anchor,
                "sentenceText": sentence,
                "feedback": feedback,
                "findings": split_feedback_findings(feedback, "SPAN"),
                "source": source,
            })
    if not results:
        results.append({
            "index": 1,
            "anchorText": None,
            "sentenceText": None,
            "feedback": text,
            "findings": split_feedback_findings(text, "SENTENCE"),
            "source": source,
        })
    return results


def parse_rewrite(
    text: str | None,
    source: str,
    provenance: dict[str, Any] | None = None,
) -> dict[str, Any] | None:
    if not text:
        return None
    layers: list[str] = []
    note = None
    rewritten = text
    prefix = re.match(r"^\[([^\]]+)\]\s*(.*)$", text, flags=re.S)
    if prefix:
        metadata = prefix.group(1)
        rewritten = prefix.group(2).strip()
        layer_part = metadata.split("|", 1)[0]
        layer_part = re.sub(r"^(层|layer)\s*:\s*", "", layer_part, flags=re.I)
        layers = unique([part.strip().upper() for part in re.split(r"[+,/，、]", layer_part) if part.strip()])
        if "|" in metadata:
            note = clean(metadata.split("|", 1)[1])
    if "→" in rewritten:
        left, right = rewritten.split("→", 1)
        note = clean(left) or note
        rewritten = right.strip()
    return {
        "source": source,
        "layers": layers,
        "text": clean(rewritten),
        "note": note,
        "raw": text,
        "provenance": provenance,
    }


def parse_global(rows: list[list[str | None]]) -> tuple[str | None, list[dict[str, Any]], list[dict[str, Any]]]:
    question = None
    dimensions = []
    other = []
    for excel_row, row in enumerate(rows[1:], start=2):
        label = (row[0] or "").strip() if row else ""
        value = row[1] if len(row) > 1 else None
        key = label.upper()
        source = source_cell(SHEET_ALIASES["global"], excel_row, "B")
        if label in {"完整原题", "原题", "题目", "QUESTION"}:
            question = value
        elif key in DIMENSIONS or key == "OVERALL":
            dimensions.append({
                "dimension": key,
                "score": parse_band(value),
                "feedback": value,
                "evidenceSids": extract_sid_refs(value),
                "findings": split_feedback_findings(value, "ESSAY"),
                "source": source,
            })
        else:
            other.append({
                "label": label,
                "content": value,
                "evidenceSids": extract_sid_refs(value),
                "findings": split_feedback_findings(value, "ESSAY"),
                "__sourceSheet": SHEET_ALIASES["global"],
                "__sourceRow": excel_row,
                "__sourceColumn": "B",
            })
    return question, dimensions, other


def paragraph_ranges(rows: list[list[str | None]]) -> list[dict[str, Any]]:
    result = []
    for index, row in enumerate(rows[1:]):
        excel_row = index + 2
        label = row[0] if row else None
        refs = extract_sid_refs(label)
        numbers = [number for ref in refs if (number := sid_number(ref)) is not None]
        versions = unique([sid_version(ref) for ref in refs if sid_version(ref)])
        result.append({
            "index": index,
            "label": label,
            "startSentenceNumber": min(numbers) if numbers else None,
            "endSentenceNumber": max(numbers) if numbers else None,
            "sidRefs": refs,
            "versionIds": versions,
            "source": source_cell(SHEET_ALIASES["paragraph"], excel_row, "A"),
            "dimensions": [
                {
                    "dimension": dimension,
                    "feedback": row[dim_index + 1] if len(row) > dim_index + 1 else None,
                    "evidenceSids": extract_sid_refs(row[dim_index + 1] if len(row) > dim_index + 1 else None),
                    "findings": split_feedback_findings(
                        row[dim_index + 1] if len(row) > dim_index + 1 else None,
                        "PARAGRAPH",
                    ),
                    "source": source_cell(
                        SHEET_ALIASES["paragraph"],
                        excel_row,
                        chr(ord("B") + dim_index),
                    ),
                }
                for dim_index, dimension in enumerate(DIMENSIONS)
            ],
        })
    return result


def paragraph_for_sentence(sid: str, paragraphs: list[dict[str, Any]]) -> int | None:
    number = sid_number(sid)
    if number is None:
        return None
    version = sid_version(sid)
    for paragraph in paragraphs:
        start = paragraph["startSentenceNumber"]
        end = paragraph["endSentenceNumber"]
        versions = paragraph.get("versionIds") or []
        if versions and version not in versions:
            continue
        if start is not None and end is not None and start <= number <= end:
            return paragraph["index"]
    return None


def infer_subtype(question: str | None) -> str | None:
    sample = (question or "").lower()
    if "discuss both" in sample:
        return "DISCUSSION"
    question_parts = len(re.findall(r"[?？]", sample))
    if question_parts >= 2 and re.search(r"\b(?:why|how|what|who)\b", sample):
        return "REPORT_TWO_QUESTIONS"
    if (
        "advantages outweigh" in sample
        or "outweigh the disadvantages" in sample
        or "outweigh" in sample
        or "positive or negative development" in sample
        or "advantages and disadvantages" in sample
        or "benefits of" in sample and "drawbacks" in sample
    ):
        return "ADVANTAGES_DISADVANTAGES"
    if "to what extent" in sample or "how far do you agree" in sample or re.search(r"\bdo you agree or disagree\b", sample):
        return "OPINION"
    if re.search(r"\b(?:why|how|what|who)\b", sample):
        return "REPORT"
    return None


def parse_workbook(path: Path, holdout_rate: float, question_overrides: dict[str, Any]) -> dict[str, Any]:
    raw_bytes = path.read_bytes()
    file_hash = sha256_bytes(raw_bytes)
    workbook = load_workbook(path, read_only=True, data_only=False)
    raw_sheets = {key: rows_for_sheet(workbook, name) for key, name in SHEET_ALIASES.items()}
    workbook.close()

    warnings: list[str] = []
    missing_sheets = [name for key, name in SHEET_ALIASES.items() if not raw_sheets[key]]
    if missing_sheets:
        warnings.append("missing_sheets=" + ",".join(missing_sheets))

    question, global_dimensions, global_other = parse_global(raw_sheets["global"])
    question_source = "WORKBOOK_GLOBAL" if question else None
    question_source_key = None
    override = question_overrides.get(path.name)
    if override:
        question = clean(override.get("questionText"))
        question_source = override.get("sourceType", "CURATED_OVERRIDE")
        question_source_key = override.get("sourceKey")
    paragraphs = paragraph_ranges(raw_sheets["paragraph"])
    sentence_rows = raw_sheets["sentence"][1:] if raw_sheets["sentence"] else []
    sentences: list[dict[str, Any]] = []
    sid_counts: Counter[str] = Counter()
    for index, row in enumerate(sentence_rows):
        excel_row = index + 2
        if len(row) < 2 or not row[0] or not row[1]:
            continue
        sid = row[0].strip().lower()
        sid_counts[sid] += 1
        text = row[1]
        dimensions = parse_dimension_feedback(
            row[3] if len(row) > 3 else None,
            source_cell(SHEET_ALIASES["sentence"], excel_row, "D"),
            "SENTENCE",
        )
        if len(dimensions) != 4:
            warnings.append(f"{sid}:dimension_count={len(dimensions)}")
        sentences.append({
            "sid": sid,
            "versionId": sid_version(sid) or "default",
            "sidOccurrence": sid_counts[sid],
            "index": index,
            "paragraphIndex": paragraph_for_sentence(sid, paragraphs),
            "text": text,
            "source": source_cell(SHEET_ALIASES["sentence"], excel_row, "B"),
            "wordAnnotations": parse_word_annotations(
                row[2] if len(row) > 2 else None,
                source_cell(SHEET_ALIASES["sentence"], excel_row, "C"),
            ),
            "dimensions": dimensions,
            "dimensionFeedbackRaw": row[3] if len(row) > 3 else None,
            "aiRewrite": parse_rewrite(
                row[4] if len(row) > 4 else None,
                "AI",
                source_cell(SHEET_ALIASES["sentence"], excel_row, "E"),
            ),
            "teacherRewrite": parse_rewrite(
                row[5] if len(row) > 5 else None,
                "TEACHER",
                source_cell(SHEET_ALIASES["sentence"], excel_row, "F"),
            ),
        })
    duplicates = sorted(sid for sid, count in sid_counts.items() if count > 1)
    if duplicates:
        warnings.append("duplicate_sids=" + ",".join(duplicates))

    families = sorted(set(sentence["versionId"] for sentence in sentences))
    multi_essay = len(families) > 1
    if multi_essay:
        warnings.append("multiple_sid_families=" + ",".join(families))

    grouped: dict[int, list[dict[str, Any]]] = defaultdict(list)
    unassigned = []
    for sentence in sentences:
        paragraph_index = sentence["paragraphIndex"]
        if paragraph_index is None:
            unassigned.append(sentence)
        else:
            grouped[paragraph_index].append(sentence)
    if unassigned:
        warnings.append(f"unassigned_sentences={len(unassigned)}")
        fallback_index = len(paragraphs)
        paragraphs.append({
            "index": fallback_index,
            "label": "未映射句子",
            "startSentenceNumber": None,
            "endSentenceNumber": None,
            "dimensions": [],
            "sidRefs": [],
            "versionIds": [],
        })
        grouped[fallback_index] = unassigned
        for sentence in unassigned:
            sentence["paragraphIndex"] = fallback_index

    essay_parts = []
    cursor = 0
    for paragraph in paragraphs:
        paragraph_sentences = grouped.get(paragraph["index"], [])
        paragraph_text_parts = []
        if essay_parts:
            cursor += 2
        paragraph_start = cursor
        for sentence_index, sentence in enumerate(paragraph_sentences):
            if sentence_index:
                cursor += 1
            sentence["startOffset"] = cursor
            sentence["endOffset"] = cursor + len(sentence["text"])
            paragraph_text_parts.append(sentence["text"])
            cursor = sentence["endOffset"]
        paragraph["text"] = " ".join(paragraph_text_parts)
        paragraph["startOffset"] = paragraph_start
        paragraph["endOffset"] = cursor
        essay_parts.append(paragraph["text"])
    essay_text = "\n\n".join(essay_parts).strip()

    source_paragraphs = row_objects(raw_sheets["essay_paragraphs"], SHEET_ALIASES["essay_paragraphs"])
    for index, source in enumerate(source_paragraphs):
        source_text = next((value for key, value in source.items() if key != "标注" and value), None)
        if index < len(paragraphs) and source_text and paragraphs[index]["text"]:
            ratio = SequenceMatcher(None, re.sub(r"\s+", " ", source_text), re.sub(r"\s+", " ", paragraphs[index]["text"])).ratio()
            paragraphs[index]["sourceTextSimilarity"] = round(ratio, 4)
            if ratio < 0.80:
                warnings.append(f"paragraph_{index}:source_text_similarity={ratio:.3f}")

    word_count = english_word_count(essay_text)
    if not question:
        warnings.append("missing_question")
    if word_count < 180:
        warnings.append(f"short_essay={word_count}")
    if len(sentences) < 5:
        warnings.append(f"few_sentences={len(sentences)}")

    fatal_for_rag = multi_essay or bool(duplicates) or not question or word_count < 150 or len(sentences) < 5
    split_bucket = int(file_hash[:8], 16) / 0xFFFFFFFF
    split = "holdout" if split_bucket < holdout_rate else "train"
    allowed_for_rag = not fatal_for_rag and split == "train"

    model_rows = row_objects(raw_sheets["model"], SHEET_ALIASES["model"])
    model_text = "\n\n".join(
        value
        for row in model_rows
        for key, value in row.items()
        if not key.startswith("__")
        and isinstance(value, str)
        and value
        and ("范文" in key or "essay" in key.lower())
        and not any(marker in value for marker in ("缺失", "没有标注", "未提供"))
    )
    model_status = "COMPLETE" if english_word_count(model_text) >= 180 else "MISSING_OR_PARTIAL"

    return {
        "documentKey": file_hash[:24],
        "fileName": path.name,
        "sourcePath": str(path.resolve()),
        "fileHash": file_hash,
        "textHash": sha256_text(essay_text),
        "task": "TASK2",
        "subtype": infer_subtype(question),
        "questionText": question,
        "questionSource": question_source,
        "questionSourceKey": question_source_key,
        "essayText": essay_text,
        "wordCount": word_count,
        "overallBand": next((item["score"] for item in global_dimensions if item["dimension"] == "OVERALL"), None),
        "globalAssessments": global_dimensions,
        "globalOther": global_other,
        "paragraphs": paragraphs,
        "sentences": sentences,
        "opinions": row_objects(raw_sheets["opinions"], SHEET_ALIASES["opinions"]),
        "teachingNotes": row_objects(raw_sheets["teaching"], SHEET_ALIASES["teaching"]),
        "expansions": row_objects(raw_sheets["expansions"], SHEET_ALIASES["expansions"]),
        "modelEssay": {"status": model_status, "wordCount": english_word_count(model_text), "text": clean(model_text), "rows": model_rows},
        "modelAnalysis": row_objects(raw_sheets["model_analysis"], SHEET_ALIASES["model_analysis"]),
        "sourceParagraphRows": source_paragraphs,
        "quality": {
            "status": "NEEDS_REVIEW" if fatal_for_rag else "STRUCTURALLY_VALID",
            "multipleEssaysSuspected": multi_essay,
            "warnings": unique(warnings),
            "allowedForRag": allowed_for_rag,
            "excludeFromEval": split == "train",
            "split": split,
            "labelAuthority": "AI_ASSISTED_OR_UNKNOWN",
        },
        "stats": {
            "paragraphs": len(paragraphs),
            "sentences": len(sentences),
            "sentenceDimensionAssessments": sum(len(sentence["dimensions"]) for sentence in sentences),
            "wordAnnotations": sum(len(sentence["wordAnnotations"]) for sentence in sentences),
            "paragraphDimensionAssessments": sum(len(paragraph["dimensions"]) for paragraph in paragraphs),
            "globalAssessments": len(global_dimensions),
            "assessmentFindings": (
                sum(len(item.get("findings", [])) for item in global_dimensions)
                + sum(
                    len(item.get("findings", []))
                    for paragraph in paragraphs
                    for item in paragraph["dimensions"]
                )
                + sum(
                    len(item.get("findings", []))
                    for sentence in sentences
                    for item in sentence["dimensions"]
                )
                + sum(
                    len(item.get("findings", []))
                    for sentence in sentences
                    for item in sentence["wordAnnotations"]
                )
            ),
            "explicitEvidenceRefs": (
                sum(len(item.get("evidenceSids", [])) for item in global_dimensions)
                + sum(
                    len(item.get("evidenceSids", []))
                    for paragraph in paragraphs
                    for item in paragraph["dimensions"]
                )
                + sum(
                    len(item.get("evidenceSids", []))
                    for sentence in sentences
                    for item in sentence["dimensions"]
                )
            ),
        },
    }


def main() -> None:
    args = parse_args()
    question_overrides = {}
    if args.question_overrides:
        override_payload = json.loads(args.question_overrides.read_text(encoding="utf-8"))
        question_overrides = override_payload.get("documents", override_payload)
    input_dir = args.input_dir.resolve()
    output = args.output.resolve()
    documents = []
    errors = []
    for path in sorted(input_dir.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        try:
            documents.append(parse_workbook(path, args.holdout_rate, question_overrides))
        except Exception as exc:
            errors.append({"sourcePath": str(path.resolve()), "error": f"{type(exc).__name__}: {exc}"})

    manifest = {
        "schemaVersion": "rag-excel-task2-manifest/2.0.0",
        "parserVersion": PARSER_VERSION,
        "generatedAt": datetime.now().astimezone().isoformat(),
        "inputDirectory": str(input_dir),
        "summary": {
            "files": len(documents) + len(errors),
            "documents": len(documents),
            "errors": len(errors),
            "structurallyValid": sum(d["quality"]["status"] == "STRUCTURALLY_VALID" for d in documents),
            "needsReview": sum(d["quality"]["status"] == "NEEDS_REVIEW" for d in documents),
            "train": sum(d["quality"]["split"] == "train" for d in documents),
            "holdout": sum(d["quality"]["split"] == "holdout" for d in documents),
            "allowedForRag": sum(d["quality"]["allowedForRag"] for d in documents),
            "paragraphs": sum(d["stats"]["paragraphs"] for d in documents),
            "sentences": sum(d["stats"]["sentences"] for d in documents),
            "sentenceDimensionAssessments": sum(d["stats"]["sentenceDimensionAssessments"] for d in documents),
            "wordAnnotations": sum(d["stats"]["wordAnnotations"] for d in documents),
            "paragraphDimensionAssessments": sum(d["stats"]["paragraphDimensionAssessments"] for d in documents),
            "globalAssessments": sum(d["stats"]["globalAssessments"] for d in documents),
            "assessmentFindings": sum(d["stats"]["assessmentFindings"] for d in documents),
            "explicitEvidenceRefs": sum(d["stats"]["explicitEvidenceRefs"] for d in documents),
            "completeModelEssays": sum(d["modelEssay"]["status"] == "COMPLETE" for d in documents),
        },
        "qualityQueue": [
            {"fileName": d["fileName"], "status": d["quality"]["status"], "warnings": d["quality"]["warnings"]}
            for d in documents
            if d["quality"]["status"] == "NEEDS_REVIEW" or d["quality"]["warnings"]
        ],
        "errors": errors,
        "documents": documents,
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    console = {"summary": manifest["summary"], "errors": errors}
    if not args.summary_only:
        console["qualityQueue"] = manifest["qualityQueue"]
    print(json.dumps(console, ensure_ascii=False, indent=2))
    print(f"Manifest: {output}")


if __name__ == "__main__":
    main()
