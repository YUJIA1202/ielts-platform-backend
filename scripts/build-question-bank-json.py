import datetime as dt
import json
import re
import sys
from pathlib import Path

import openpyxl


REQUIRED_COLUMNS = [
    "#",
    "日期",
    "形式",
    "地区",
    "话题大类",
    "话题二级",
    "题型",
    "相似题",
    "题目原文",
    "来源",
]

TASK2_SUBTYPE_ALIASES = {
    "程度": "同意与否/程度同意",
    "程度同意": "同意与否/程度同意",
    "同意与否": "同意与否/程度同意",
    "双边": "双边/讨论双方",
    "讨论双方": "双边/讨论双方",
    "报告": "报告/回答两个问题",
    "回答两个问题": "报告/回答两个问题",
    "优缺点": "优缺点/积极消极",
    "积极消极": "优缺点/积极消极",
    "其优缺点": "优缺点/积极消极",
    "其他优缺点": "优缺点/积极消极",
}


def clean(value):
    return "" if value is None else str(value).strip()


def normalize_subtype(value):
    subtype = clean(value)
    return TASK2_SUBTYPE_ALIASES.get(subtype, subtype)


def normalize_region(value):
    region = clean(value)
    return "" if region in {"1", "默认", "普通"} else region


def parse_date(value):
    if value in (None, ""):
        return None, None, None

    if isinstance(value, (dt.date, dt.datetime)):
        return value.strftime("%Y-%m-%d"), value.year, value.month

    text = clean(value)
    match = re.match(r"^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})$", text)
    if match:
        year, month, day = map(int, match.groups())
        return f"{year:04d}-{month:02d}-{day:02d}", year, month

    year_match = re.search(r"(19\d{2}|20\d{2})", text)
    return None, int(year_match.group(1)) if year_match else None, None


def main():
    if len(sys.argv) != 3:
        raise SystemExit("Usage: python scripts/build-question-bank-json.py <input.xlsx> <output.json>")

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    workbook = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    if "题库" not in workbook.sheetnames:
        raise SystemExit("Workbook must contain a sheet named 题库")

    sheet = workbook["题库"]
    rows = list(sheet.iter_rows(values_only=True))
    header = [clean(cell) for cell in rows[0]]
    column_index = {name: index for index, name in enumerate(header)}

    missing = [name for name in REQUIRED_COLUMNS if name not in column_index]
    if missing:
        raise SystemExit(f"Missing columns: {missing}")

    records = []
    for row in rows[1:]:
        if not any(value not in (None, "") for value in row):
            continue

        content = clean(row[column_index["题目原文"]])
        if not content:
            continue

        try:
            source_row = int(row[column_index["#"]])
        except Exception:
            source_row = len(records) + 1

        exam_date, year, month = parse_date(row[column_index["日期"]])
        topic = clean(row[column_index["话题大类"]])

        records.append(
            {
                "sourceKey": f"ielts-task2-final-revised:{source_row}",
                "sourceRow": source_row,
                "task": "TASK2",
                "subtype": normalize_subtype(row[column_index["题型"]]),
                "topic": topic,
                "topicCategory": topic,
                "topicSubcategory": clean(row[column_index["话题二级"]]),
                "content": content,
                "source": clean(row[column_index["来源"]]),
                "examDate": exam_date,
                "year": year,
                "month": month,
                "testMode": clean(row[column_index["形式"]]),
        "region": normalize_region(row[column_index["地区"]]),
                "similarGroup": clean(row[column_index["相似题"]]),
            }
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(records, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(f"Question bank JSON written: {output_path}")
    print(f"records={len(records)} maxSourceRow={max(record['sourceRow'] for record in records)}")


if __name__ == "__main__":
    main()
